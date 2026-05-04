import requests
import json
import re
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import os

# ========================================================
# 1. 系統與實驗參數設定 (System & Experiment Configurations)
# ========================================================
OLLAMA_API = "http://localhost:11434/api/generate"
MODEL_L1 = "my-soc-agent-en"     # L1 意圖萃取代理人
MODEL_L3 = "llama3.1"            # L3 嚴厲語意裁判

# 🎯 指定高難度的 50 題綜合考卷
DATASET_FILE = "complex_apt_dataset_50.jsonl"  
EXPORT_EXCEL = "benchmark_results_massive.xlsx"

MAX_TEST_CASES = 30    # 設為 0 代表全部跑完
START_OFFSET = 10      

# ========================================================
# 2. 核心功能函數與防護網 (Core Functions & Safeguards)
# ========================================================
def call_llm(prompt, model, is_json=True):
    payload = {"model": model, "prompt": prompt, "stream": False, "options": {"temperature": 0.0}}
    if is_json: payload["format"] = "json"
    try:
        return requests.post(OLLAMA_API, json=payload).json().get("response", "").strip()
    except Exception as e:
        return ""

def extract_json(text):
    try:
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match: return json.loads(match.group())
    except: pass
    return {}

def sanitize_for_excel(value):
    if isinstance(value, str):
        illegal_chars = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        return illegal_chars.sub(r'', value)
    return value

# ========================================================
# 3. 初始化 RAG 引擎與 Excel 報表
# ========================================================
print("⏳ 正在啟動 RAG 向量引擎與載入 MITRE 知識庫...")
os.environ["HF_TOKEN"] = "" # 若有 Hugging Face Token 可填入以消除警告
embedder = SentenceTransformer('all-MiniLM-L6-v2')

try:
    with open("mitre_knowledge_base.json", "r", encoding="utf-8") as f:
        mitre_kb = json.load(f)
    mitre_ids, mitre_descriptions = list(mitre_kb.keys()), list(mitre_kb.values())
    kb_embeddings = embedder.encode(mitre_descriptions)
except FileNotFoundError:
    print("❌ 找不到 mitre_knowledge_base.json，請確認知識庫存在！")
    exit()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Benchmark Results"
headers = ["Task ID", "Source File", "Ground Truth (IDs)", "L1 Intent", "RAG Prediction", "EM Match", "L3 Score", "Judge Reason"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

processed_tasks = 0      
success_count = 0        
exact_match_count = 0    
failed_cases = []        

# ========================================================
# 4. 啟動評測管線 
# ========================================================
target_text = f"{MAX_TEST_CASES} 筆" if MAX_TEST_CASES > 0 else "全部"
print(f"\n🚀 開始執行【LLM-as-a-Judge 語意等效評估】 (目標: {target_text})\n" + "-"*60)

with open(DATASET_FILE, 'r', encoding='utf-8') as f:
    for line_idx, line in enumerate(f):
        if line_idx < START_OFFSET: continue
        processed_tasks += 1
        
        d = json.loads(line)
        valid_ttps = d.get('valid_ttps', [])
        source_file = d.get('source_file', 'Unknown')
        
        log_match = re.search(r'### Input:\n(.*?)\n### Response:', d['text'], re.DOTALL)
        input_log = log_match.group(1) if log_match else d['text']

        # --- [A. L1 代理人萃取意圖 (攻防平衡 Few-Shot 版)] ---
        l1_prompt = f"""### Instruction:
You are an elite threat hunter analyzing endpoint logs. 
CRITICAL RULE: Not all logs are malicious. Many are just routine background noise. If the log appears benign, normal, or ambiguous, you MUST state the intent as "Normal System Behavior / Unknown Threat".

[Examples]
Log: "procdump.exe -ma lsass.exe" -> Intent: "Adversary is attempting OS Credential Dumping."
Log: "ping.exe 8.8.8.8" -> Intent: "Normal System Behavior / Unknown Threat. Routine network check."
Log: "wevtutil cl Security" -> Intent: "Adversary is attempting Defense Evasion by clearing event logs."
Log: "svchost.exe -k LocalService" -> Intent: "Normal System Behavior / Unknown Threat. Routine Windows service execution."

Now analyze the following log. Output ONLY valid JSON: {{"Action": "what it literally does", "Intent": "the tactical goal or 'Normal System Behavior / Unknown Threat'"}}.
### Input:
{input_log}
### Response:"""

	# ⚠️ 就是下面這兩行剛才不小心被刪掉了，它們負責取得 LLM 的回答！
        l1_res = extract_json(call_llm(l1_prompt, MODEL_L1))
        intent = str(l1_res.get("Intent", "Unknown"))
        

        # --- [B. 動態閾值 RAG 判定戰術 (提高準度)] ---
        query_embedding = embedder.encode([intent])
        sims = cosine_similarity(query_embedding, kb_embeddings)[0]
        best_idx = np.argmax(sims)
        
        # 動態門檻：正常行為攔截門檻 0.60，異常行為的最低門檻從 0.35 提高到 0.45，避免亂猜
        is_normal = any(k in intent.lower() for k in ["normal", "benign", "standard", "legitimate", "unknown"])
        rag_ttp = mitre_ids[best_idx] if (sims[best_idx] > (0.60 if is_normal else 0.45)) else "Unknown Threat"

        # --- [C. 傳統 EM 判定與短路邏輯 (Short-circuit)] ---
        is_exact = rag_ttp in valid_ttps
        
        if is_exact:
            # 🌟 系統優化：如果是 100% 精準命中，直接給滿分，不需要浪費時間問裁判！
            exact_match_count += 1
            score = 2
            reason = "Exact Match. (Auto-approved by system without LLM Judge)"
            success_count += 1
            
        else:
            # --- [D. L3 鐵面裁判 (Chain of Thought 防幻覺終極版)] ---
            truth_descriptions = []
            for ttp in valid_ttps:
                if ttp in mitre_kb:
                    truth_descriptions.append(f"{ttp}: {mitre_kb[ttp]}")
                else:
                    truth_descriptions.append(f"{ttp}: (Normal Behavior or Unknown Threat)")
            truth_desc_str = " | ".join(truth_descriptions)
            
            predicted_desc = mitre_kb.get(rag_ttp, "Normal Behavior or Unknown Threat")

            judge_prompt = f"""You are an extremely strict, unforgiving SOC Auditor evaluating a junior analyst.

[Ground Truth]: {valid_ttps} - {truth_desc_str}
[Analyst Prediction]: {rag_ttp} - {predicted_desc}
[Analyst Intent Description]: {intent}

[Evaluation Rules]
1. Compare the core tactical behavior strictly based on the provided definitions.
2. If the MITRE Tactics are fundamentally different (e.g., 'Credential Access' vs 'Network/Command and Control', or 'Unknown' vs an actual attack), they DO NOT MATCH. You MUST score 0.
3. If the core tactical objective is identical despite different IDs (e.g., both are clearly 'Credential Dumping' or both are 'Privilege Escalation'), score 1.
4. Do NOT make up vague connections like "both are malicious" or "both access information".

Output ONLY valid JSON. You MUST output "Reason" FIRST, then "Score" (1 for partial tactical match, 0 for fail).
{{
    "Reason": "Analyze the technical alignment or difference step-by-step.",
    "Score": number
}}"""
            
            judge_res = extract_json(call_llm(judge_prompt, MODEL_L3))
            
            raw_score = judge_res.get("Score", 0)
            try:
                score = int(list(raw_score.values())[0]) if isinstance(raw_score, dict) else int(raw_score)
            except (TypeError, ValueError):
                score = 0
                
            reason = str(judge_res.get("Reason", "Format Error"))
            
            if score >= 1: 
                success_count += 1
            else:
                failed_cases.append({
                    "id": line_idx + 1,
                    "source": source_file,
                    "intent": intent,
                    "rag_ttp": rag_ttp,
                    "truth": valid_ttps,
                    "reason": reason
                })

        # --- [E. 寫入 Excel 與終端機輸出] ---
        row_data = [line_idx + 1, source_file, str(valid_ttps), intent, rag_ttp, is_exact, score, reason]
        clean_row_data = [sanitize_for_excel(cell) for cell in row_data]
        ws.append(clean_row_data)
        
        if score == 0: 
            for cell in ws[ws.max_row]: cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        print(f"➤ Task_{line_idx + 1:03d} | L3 Score: {score} | EM: {is_exact} | 預測: {rag_ttp}")
        
        if MAX_TEST_CASES > 0 and processed_tasks >= MAX_TEST_CASES: break

# ========================================================
# 5. 輸出最終數據與終端機總結報告
# ========================================================
wb.save(EXPORT_EXCEL)
em_rate = (exact_match_count / processed_tasks) * 100 if processed_tasks > 0 else 0
l3_rate = (success_count / processed_tasks) * 100 if processed_tasks > 0 else 0

print("\n" + "="*60)
print(f"📊 實驗數據總結已儲存至：{EXPORT_EXCEL}")
print("="*60)
print(f"➤ 實際測試樣本數：{processed_tasks} 筆")
print(f"➤ 成功筆數 (Score 1-2)：{success_count} 筆")
print(f"➤ 失敗筆數 (Score 0)  ：{len(failed_cases)} 筆")
print(f"➤ 傳統字串嚴格命中率 (Exact Match): {em_rate:.1f}%")
print(f"➤ LLM 裁判實務有效告警率 (L3 Judge): {l3_rate:.1f}%")
print("="*60)

if failed_cases:
    print("\n🚨 失敗任務細節與原因分析 (Score = 0)")
    print("-" * 60)
    for case in failed_cases:
        print(f"❌ Task_{case['id']:03d} | 來源: {case['source']}")
        print(f"   [L1 意圖] {case['intent'][:80]}...") 
        print(f"   [系統預測] {case['rag_ttp']} | [真實解答] {case['truth']}")
        print(f"   [裁判理由] {case['reason']}")
        print("-" * 60)
else:
    print("\n🎉 太棒了！本次測試沒有任何失敗任務 (Score=0)，系統表現完美！")
    print("=" * 60)